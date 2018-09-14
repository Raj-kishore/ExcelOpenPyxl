import os
from openpyxl import load_workbook
# set the cd path to access excel with openpyxl
os.chdir("C:\\Users\\raj\\Desktop\\SAPaskmesearch v2")
print("Current path is "+os.getcwd())
# Copy and Create a 2nd excel sheet manually and add the column on 9th place horizontally that will take the modified values from the 6th column
#load 2nd excel sheet
print("Loading the excel sheet...please wait")
wb2 = load_workbook('AskMe_SearchResults_Updated_Copy.xlsx')
# wait to load sheetnames
print (wb2.sheetnames)
# assign sheets to a list
sl2 = wb2.sheetnames
sl12 = sl2[1]
print("sheet name "+sl12)
# create an object for 2md sheet of 2nd excel file
worksheet2 = wb2[sl12]
count = 0
for row in range(1, worksheet2.max_row):
    count+=1
    #rows object for 9th column
    cell = worksheet2.cell(row=row, column=9)
    #rows object for 6th column
    cellf = worksheet2.cell(row=row, column=6)
    #take value of each row from 6th column
    string = cellf.value
    print(str(count) +" : "+ string),
    # None types should be avoided
    if string != None:
        # substitution code
        replacedString = ''
        countChRep = 0
        for b in string:
            if b != string[0] or b != string[len(string)-1]:
                #do replacement
                if b == '-' or b == '_' or b == ' ':
                    countChRep +=1
                    if countChRep <=1:
                        b = ','
                    else:
                        b = ''
                    replacedString += b
                else:
                    replacedString += b
                    countChRep = 0
        print(">>"+replacedString+"<<")
    # time to assign
    #if cell.value is not None:
    cell.value = replacedString
# assign it to 0 for next row
countChRepetion = 0
# finally save it.
wb2.save('AskMe_SearchResults_Updated_Copy.xlsx')
# Done
'''© Rajkishor, Dt : 12-Sep-2018'''